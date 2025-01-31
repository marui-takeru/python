import openpyxl

# エクセルファイルを読み込む
wb = openpyxl.load_workbook('時間スケール調査用.xlsx')

# アクティブなシートを選択（必要に応じてシート名を指定してください）
sheet = wb.active

# データの開始行を指定（6行目からデータが始まる）
data_start_row = 6

# 必要なデータの範囲を計算（例: 入力で指定された秒間のデータを取得）
duration_seconds = float(input("平均したい秒: "))  # ユーザーから取得したい時間（秒）
data_rows = int(duration_seconds / 0.1)  # 各行が0.1秒を表すため、行数を計算
end_row = data_rows + data_start_row

# データのヘッダーを設定
sheet['I5'] = 'U'  # I列のヘッダーを U に設定
sheet['J5'] = 'V'  # J列のヘッダーを V に設定
sheet['K5'] = 'W'  # K列のヘッダーを W に設定
sheet['N5'] = "U' * W' average"  # N列のヘッダーを U'*W' average に設定

# C列の値をI列にコピーし、D列とB列の値に負の数をかけてそれぞれJ列とK列にコピー
for row in range(data_start_row, sheet.max_row + 1):  # 必要な行の範囲で処理
    # C列の値を取得し、I列にコピー
    value_c = sheet[f'C{row}'].value
    sheet[f'I{row}'] = value_c

    # D列の値に負の数をかけ、J列にコピー
    value_d = sheet[f'D{row}'].value
    if isinstance(value_d, (int, float)):
        sheet[f'J{row}'] = -value_d

    # B列の値に負の数をかけ、K列にコピー
    value_b = sheet[f'B{row}'].value
    if isinstance(value_b, (int, float)):
        sheet[f'K{row}'] = -value_b

# 指定された範囲でスライディングウィンドウ平均を計算
for start_row in range(data_start_row, sheet.max_row - data_rows + 1):
    average_row = int((data_rows) / 2 + data_start_row + 1)  # 中央付近の行

    # 平均値の計算に使う中心行からの平均値を取得
    U_average = sheet[f'I{average_row}'].value  # I列の平均値
    W_average = sheet[f'K{average_row}'].value  # K列の平均値
    V_average = sheet[f'J{average_row}'].value  # J列の平均値

    # U' * W' のサマリーを計算
    sum_uw = 0

    # 各列の値を平均値との差分としてコピーし、U' * W' を計算
    for row in range(data_start_row, end_row + 1):
        # I列の値を取得し、 U' を設定
        value_i = sheet[f'I{row}'].value
        if isinstance(value_i, (int, float)):
            UU = value_i - U_average

        # K列の値を取得し、 W' を設定
        value_k = sheet[f'K{row}'].value
        if isinstance(value_k, (int, float)):
            WW = value_k - W_average

        # J列の値を取得し、 V' を設定
        value_j = sheet[f'J{row}'].value
        if isinstance(value_j, (int, float)):
            VV = value_j - V_average

        # U' * W' を計算し、N列に設定
        if isinstance(UU, (int, float)) and isinstance(WW, (int, float)):
            value_uw = UU * WW
            sum_uw += value_uw

    # U' * W' の平均を N列 に設定
    average_uw = sum_uw / (end_row - data_start_row + 1)
    sheet[f'N{data_start_row}'] = - average_uw

    data_start_row += 1
    end_row += 1
    print(data_start_row, end_row)

# 加工後のエクセルファイルを保存
wb.save('時間スケール調査用_加工済.xlsx')

print(f"加工が完了しました。{duration_seconds}秒間のデータを取得しました。")
